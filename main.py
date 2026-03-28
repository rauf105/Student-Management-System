import tkinter
from tkinter import messagebox, simpledialog
import openpyxl
import os

class ExcelManager:
    FILEPATH = r"D:\python_Midproject\data.xlsx"

    @staticmethod
    def save_data(data_list):
        if not os.path.exists(ExcelManager.FILEPATH):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["First Name", "Last Name", "ID", "Course Name", "Section", "Mark"])
            workbook.save(ExcelManager.FILEPATH)

        workbook = openpyxl.load_workbook(ExcelManager.FILEPATH)
        sheet = workbook.active
        sheet.append(data_list)
        workbook.save(ExcelManager.FILEPATH)

    @staticmethod
    def read_all():
        if not os.path.exists(ExcelManager.FILEPATH):
            return None
        workbook = openpyxl.load_workbook(ExcelManager.FILEPATH)
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))

    @staticmethod
    def search_by_id(student_id):
        data = ExcelManager.read_all()
        if data:
            for row in data[1:]:
                if str(row[2]) == student_id:
                    return row
        return None

    @staticmethod
    def update_mark(student_id, new_mark):
        if not os.path.exists(ExcelManager.FILEPATH):
            return False
        workbook = openpyxl.load_workbook(ExcelManager.FILEPATH)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if str(row[2].value) == student_id:
                row[5].value = new_mark
                workbook.save(ExcelManager.FILEPATH)
                return True
        return False

    @staticmethod
    def delete_by_id(student_id):
        if not os.path.exists(ExcelManager.FILEPATH):
            return False
        workbook = openpyxl.load_workbook(ExcelManager.FILEPATH)
        sheet = workbook.active
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if str(row[2].value) == student_id:
                sheet.delete_rows(i)
                workbook.save(ExcelManager.FILEPATH)
                return True
        return False

    @staticmethod
    def sort_by_mark():
        if not os.path.exists(ExcelManager.FILEPATH):
            return False
        workbook = openpyxl.load_workbook(ExcelManager.FILEPATH)
        sheet = workbook.active
        data = list(sheet.iter_rows(min_row=2, values_only=True))
        if not data:
            return False
        data.sort(key=lambda x: int(x[5]))
        sheet.delete_rows(2, sheet.max_row)
        for row in data:
            sheet.append(row)
        workbook.save(ExcelManager.FILEPATH)
        return True


class DataEntryFrame(tkinter.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.create_widgets()

    def create_widgets(self):
        # User Info
        user_info_frame = tkinter.LabelFrame(self, text="User Information")
        user_info_frame.grid(row=0, column=0, padx=20, pady=10)

        tkinter.Label(user_info_frame, text="First Name").grid(row=0, column=0)
        tkinter.Label(user_info_frame, text="Last Name").grid(row=0, column=1)
        tkinter.Label(user_info_frame, text="ID").grid(row=0, column=2)

        self.first_name_entry = tkinter.Entry(user_info_frame)
        self.last_name_entry = tkinter.Entry(user_info_frame)
        self.id_entry = tkinter.Entry(user_info_frame)

        self.first_name_entry.grid(row=1, column=0)
        self.last_name_entry.grid(row=1, column=1)
        self.id_entry.grid(row=1, column=2)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Course Info
        course_frame = tkinter.LabelFrame(self, text="Course Information")
        course_frame.grid(row=1, column=0, padx=20, pady=10)

        tkinter.Label(course_frame, text="Course Name").grid(row=0, column=0)
        tkinter.Label(course_frame, text="Section").grid(row=0, column=1)
        tkinter.Label(course_frame, text="Mark").grid(row=0, column=2)

        self.course_entry = tkinter.Entry(course_frame)
        self.section_entry = tkinter.Entry(course_frame)
        self.mark_entry = tkinter.Entry(course_frame)

        self.course_entry.grid(row=1, column=0)
        self.section_entry.grid(row=1, column=1)
        self.mark_entry.grid(row=1, column=2)

        for widget in course_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        
        terms_frame = tkinter.LabelFrame(self, text="Terms & Conditions")
        terms_frame.grid(row=2, column=0, padx=20, pady=10)

        self.accept_var = tkinter.StringVar(value="Not Accepted")
        tkinter.Checkbutton(
            terms_frame,
            text="I accept the terms and conditions.",
            variable=self.accept_var,
            onvalue="Accepted",
            offvalue="Not Accepted"
        ).grid(row=0, column=0)

        # Buttons
        tkinter.Button(self, text="Enter Data", command=self.enter_data).grid(row=3, column=0, padx=20, pady=10)
        tkinter.Button(self, text="Dashboard", command=self.master.show_dashboard).grid(row=4, column=0, padx=20, pady=10)

    def enter_data(self):
        accepted = self.accept_var.get()
        if accepted != "Accepted":
            messagebox.showwarning("Error", "You have not accepted the terms.")
            return

        firstname = self.first_name_entry.get()
        lastname = self.last_name_entry.get()
        student_id = self.id_entry.get()
        course = self.course_entry.get()
        section = self.section_entry.get()
        mark = self.mark_entry.get()

        if not all([firstname, lastname, student_id, course, section, mark]):
            messagebox.showwarning("Error", "All fields are required.")
            return

        
        if ExcelManager.search_by_id(student_id):
            messagebox.showerror("Error", f"Student ID {student_id} already exists!")
            return

        # Save data if ID is unique
        ExcelManager.save_data([firstname, lastname, student_id, course, section, mark])
        messagebox.showinfo("Success", "Data saved successfully!")

        # Clear entries
        self.first_name_entry.delete(0, 'end')
        self.last_name_entry.delete(0, 'end')
        self.id_entry.delete(0, 'end')
        self.course_entry.delete(0, 'end')
        self.section_entry.delete(0, 'end')
        self.mark_entry.delete(0, 'end')



class DashboardFrame(tkinter.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.create_widgets()

    def create_widgets(self):
        tkinter.Label(self, text="Dashboard", font=("Arial", 16)).pack(pady=10)

        tkinter.Button(self, text="See All Data", command=self.see_data).pack(pady=5)
        tkinter.Button(self, text="Search by ID", command=self.search_data).pack(pady=5)
        tkinter.Button(self, text="Update Mark", command=self.update_mark).pack(pady=5)
        tkinter.Button(self, text="Delete by ID", command=self.delete_data).pack(pady=5)
        tkinter.Button(self, text="Sort by Mark", command=self.sort_data).pack(pady=5)
        tkinter.Button(self, text="Back", command=self.master.back_to_form).pack(pady=20)

    def see_data(self):
        data = ExcelManager.read_all()
        if not data:
            messagebox.showerror("Error", "No data found.")
            return
        # Build table-like string
        header = data[0]
        all_text = f"{' | '.join(header)}\n"
        all_text += "-" * (len(all_text) + 10) + "\n"
        for row in data[1:]:
            all_text += ' | '.join(str(cell) for cell in row) + "\n"
        messagebox.showinfo("All Data", all_text)

    def search_data(self):
        student_id = simpledialog.askstring("Search", "Enter ID:")
        if not student_id:
            return
        row = ExcelManager.search_by_id(student_id)
        if row:
            messagebox.showinfo("Result", str(row))
        else:
            messagebox.showerror("Error", "ID not found")

    def update_mark(self):
        student_id = simpledialog.askstring("Update", "Enter ID to update:")
        if not student_id:
            return
        new_mark = simpledialog.askstring("Update", "Enter new mark:")
        if not new_mark:
            return
        if ExcelManager.update_mark(student_id, new_mark):
            messagebox.showinfo("Success", "Mark updated!")
        else:
            messagebox.showerror("Error", "ID not found")

    def delete_data(self):
        student_id = simpledialog.askstring("Delete", "Enter ID to delete:")
        if not student_id:
            return
        if ExcelManager.delete_by_id(student_id):
            messagebox.showinfo("Success", "Record deleted!")
        else:
            messagebox.showerror("Error", "ID not found")

    def sort_data(self):
        if ExcelManager.sort_by_mark():
            data = ExcelManager.read_all()
            if not data:
                messagebox.showerror("Error", "No data found.")
                return

            # Build table-like string
            header = data[0]
            all_text = f"{' | '.join(header)}\n"
            all_text += "-" * (len(all_text) + 10) + "\n"
            for row in data[1:]:
                all_text += ' | '.join(str(cell) for cell in row) + "\n"

            messagebox.showinfo("Sorted Data (by Mark)", all_text)
        else:
            messagebox.showerror("Error", "No data to sort.")



class App(tkinter.Tk):
    def __init__(self):
        super().__init__()
        self.title("Student Management System")
        self.geometry("600x450")

        self.data_entry_frame = DataEntryFrame(self)
        self.dashboard_frame = DashboardFrame(self)

        self.data_entry_frame.pack()

    def show_dashboard(self):
        self.data_entry_frame.pack_forget()
        self.dashboard_frame.pack()

    def back_to_form(self):
        self.dashboard_frame.pack_forget()
        self.data_entry_frame.pack()


if __name__ == "__main__":
    app = App()
    app.mainloop()